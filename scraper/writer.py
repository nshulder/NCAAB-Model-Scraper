"""
Gathers all the current days games and compiles the information into a color coded Excel document
"""

import requests
from bs4 import BeautifulSoup
from openpyxl import formatting, styles, Workbook
from openpyxl.styles import Font
from datetime import datetime
from scraper.scraper import kenpom_line
from scraper.scraper import trank_line
from scraper.utils import set_up_odds_dictionary


odds_dict = dict()


def write_game_to_file(ws, team_name, k_line, t_line, row_num):
    """

    Takes inputs for a single game and writes them out to the file

    :param ws: Excel worksheet
    :param team_name: Winning Team Name
    :param k_line: Kenpom spread
    :param t_line: Bart Torvik spread
    :param row_num: row number on worksheet
    """

    global odds_dict
    # kenpom has their favorites with a positive line, but betting convention is negative
    k_line = k_line*-1
    avg_line = (k_line + t_line) / 2
    cur_line = float(odds_dict[team_name])
    percent_delta = (cur_line - avg_line) / cur_line * 100
    ws.cell(column=1, row=row_num, value=team_name)
    ws.cell(column=2, row=row_num, value=k_line)
    ws.cell(column=3, row=row_num, value=t_line)
    ws.cell(column=4, row=row_num, value=avg_line)
    ws.cell(column=5, row=row_num, value=cur_line)
    ws.cell(column=6, row=row_num, value=percent_delta)


def getgames(browser):
    """

    Gets the current day's games and parses out the team names. Creates a timestamped excel workbook and queries Kenpom
    and Bart Torvik for their lines and writes the results out the previously created excel document.

    :param browser: Authenticated browser with access to inside information on Kenpom
    """
    global odds_dict

    page = requests.get('http://www.donbest.com/ncaab/odds/spreads/')
    soup = BeautifulSoup(page.text, 'html.parser')
    team_names = soup.find_all(class_='oddsTeamWLink')
    odds = soup.find_all(class_='oddsOpener')
    odds_dict = set_up_odds_dictionary(team_names, odds)

    # For writing to an excel sheet to easily view results
    wb = Workbook()
    filename = 'lines/{}_ncaab_lines.xlsx'.format(datetime.now().strftime('%Y%m%d'))
    ws = wb.create_sheet(title="lines")
    line_counter = 1

    # Set up headers
    header_font = Font(bold=True, size=14)
    ws.cell(column=1, row=line_counter, value="Team Name").font = header_font
    ws.cell(column=2, row=line_counter, value="Kenpom line").font = header_font
    ws.cell(column=3, row=line_counter, value="Bartorvik line").font = header_font
    ws.cell(column=4, row=line_counter, value="Avg Line").font = header_font
    ws.cell(column=5, row=line_counter, value="Opening Line").font = header_font
    ws.cell(column=6, row=line_counter, value="% Delta").font = header_font
    line_counter = line_counter + 1

    for team in team_names:
        team_name = team.contents[0]
        k_line = kenpom_line(browser, team_name)
        t_line = trank_line(team_name)
        # For now only write out games where both found a valid line
        if k_line and t_line:
            write_game_to_file(ws, team_name, k_line, t_line, line_counter)
            line_counter = line_counter + 1

    # Set up conditional formatting to make potential good bets stand out
    light_red_fill = styles.PatternFill(start_color='ffc7ce', end_color='ffc7ce', fill_type='solid')
    dark_red_fill = styles.PatternFill(start_color='FF6074', end_color='FF6074', fill_type='solid')
    light_green_fill = styles.PatternFill(start_color='B0FF99', end_color='B0FF99', fill_type='solid')
    dark_green_fill = styles.PatternFill(start_color='43EA13', end_color='43EA13', fill_type='solid')
    ws.conditional_formatting.add('F2:F100', formatting.rule.CellIsRule(operator='lessThan', formula=['-25'], fill=light_green_fill))
    ws.conditional_formatting.add('F2:F100', formatting.rule.CellIsRule(operator='lessThan', formula=['-75'], fill=dark_green_fill))
    ws.conditional_formatting.add('F2:F100', formatting.rule.CellIsRule(operator='greaterThan', formula=['25'], fill=light_red_fill))
    ws.conditional_formatting.add('F2:F100', formatting.rule.CellIsRule(operator='greaterThan', formula=['75'], fill=dark_red_fill))
    wb.save(filename=filename)
